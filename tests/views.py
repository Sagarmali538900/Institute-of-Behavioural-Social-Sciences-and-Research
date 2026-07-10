from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.contrib import messages
from .models import Exam, Section, Question, Option
from .forms import ExamForm, SectionForm, QuestionForm, OptionFormSet

import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.db import transaction
from django.http import HttpResponse, JsonResponse

@login_required
def exam_list(request):
    """
    Lists exams. Owner sees all exams; franchise users see only their own
    exams.
    """
    is_owner = request.user.is_superuser
    if is_owner:
        exams = Exam.objects.all().order_by('-created_at')
    else:
        exams = Exam.objects.filter(created_by=request.user).order_by('-created_at')
        
    return render(request, 'admin/exam_list.html', {
        'exams': exams,
        'is_owner': is_owner
    })


@login_required
def exam_create(request):
    """
    Allows Owner and franchises to create their own exams.
    """
    if request.method == 'POST':
        form = ExamForm(request.POST, user=request.user)
        if form.is_valid():
            exam = form.save(commit=False)
            exam.created_by = request.user
            exam.save()
            form.save_m2m()
            messages.success(request, f"Exam '{exam.title}' created successfully!")
            return redirect('exam_detail', exam_id=exam.id)
    else:
        form = ExamForm(user=request.user)
    return render(request, 'admin/exam_form.html', {'form': form, 'title': 'Create New Exam'})


@login_required
def exam_detail(request, exam_id):
    """
    Renders detailed structure of an exam. Franchisees can view and manage
    only the exams they created.
    """
    exam = get_object_or_404(Exam, id=exam_id)
    is_owner = request.user.is_superuser
    
    # Check permissions
    has_access = is_owner or exam.created_by == request.user
    if not has_access:
        raise PermissionDenied
 
    # Check if they have edit rights (only owner or creator can edit/delete)
    can_edit = is_owner or exam.created_by == request.user
    sections = exam.sections.all()
    
    return render(request, 'admin/exam_detail.html', {
        'exam': exam,
        'sections': sections,
        'can_edit': can_edit
    })


@login_required
def exam_edit(request, exam_id):
    """
    Edits exam settings (restricted to owner or creator).
    """
    exam = get_object_or_404(Exam, id=exam_id)
    if not request.user.is_superuser and exam.created_by != request.user:
        raise PermissionDenied
        
    if request.method == 'POST':
        form = ExamForm(request.POST, instance=exam, user=request.user)
        if form.is_valid():
            exam = form.save()
            messages.success(request, f"Exam '{exam.title}' updated successfully!")
            return redirect('exam_detail', exam_id=exam.id)
    else:
        form = ExamForm(instance=exam, user=request.user)
    return render(request, 'admin/exam_form.html', {'form': form, 'title': f"Edit Exam: {exam.title}"})


@login_required
def exam_delete(request, exam_id):
    """
    Deletes an exam (restricted to owner or creator).
    """
    exam = get_object_or_404(Exam, id=exam_id)
    if not request.user.is_superuser and exam.created_by != request.user:
        raise PermissionDenied
        
    if request.method == 'POST':
        title = exam.title
        exam.delete()
        messages.success(request, f"Exam '{title}' has been deleted.")
        return redirect('exam_list')
    return render(request, 'admin/confirm_delete.html', {'object': exam, 'back_url': redirect('exam_detail', exam_id=exam.id).url})


@login_required
def section_add(request, exam_id):
    """
    Adds a section (restricted to owner or creator).
    """
    exam = get_object_or_404(Exam, id=exam_id)
    if not request.user.is_superuser and exam.created_by != request.user:
        raise PermissionDenied
        
    if request.method == 'POST':
        form = SectionForm(request.POST)
        if form.is_valid():
            section = form.save(commit=False)
            section.exam = exam
            section.save()
            messages.success(request, f"Section '{section.name}' added successfully!")
            return redirect('exam_detail', exam_id=exam.id)
    else:
        initial_order = exam.sections.count() + 1
        form = SectionForm(initial={
            'order': initial_order,
            'duration_minutes': 10,
            'duration_seconds': 0
        })
    return render(request, 'admin/section_form.html', {'form': form, 'exam': exam, 'title': f"Add Section to {exam.title}"})


@login_required
def section_edit(request, section_id):
    """
    Edits a section (restricted to owner or creator).
    """
    section = get_object_or_404(Section, id=section_id)
    exam = section.exam
    if not request.user.is_superuser and exam.created_by != request.user:
        raise PermissionDenied
        
    if request.method == 'POST':
        form = SectionForm(request.POST, instance=section)
        if form.is_valid():
            section = form.save()
            messages.success(request, f"Section '{section.name}' updated successfully!")
            return redirect('exam_detail', exam_id=exam.id)
    else:
        form = SectionForm(instance=section)
    return render(request, 'admin/section_form.html', {'form': form, 'exam': exam, 'title': f"Edit Section: {section.name}"})


@login_required
def section_delete(request, section_id):
    """
    Deletes a section (restricted to owner or creator).
    """
    section = get_object_or_404(Section, id=section_id)
    exam = section.exam
    if not request.user.is_superuser and exam.created_by != request.user:
        raise PermissionDenied
        
    if request.method == 'POST':
        name = section.name
        section.delete()
        messages.success(request, f"Section '{name}' has been deleted.")
        return redirect('exam_detail', exam_id=exam.id)
    return render(request, 'admin/confirm_delete.html', {'object': section, 'back_url': redirect('exam_detail', exam_id=exam.id).url})


@login_required
def question_add(request, section_id):
    """
    Adds a question (restricted to owner or creator).
    """
    section = get_object_or_404(Section, id=section_id)
    exam = section.exam
    if not request.user.is_superuser and exam.created_by != request.user:
        raise PermissionDenied
        
    if request.method == 'POST':
        form = QuestionForm(request.POST, request.FILES)
        if form.is_valid():
            question = form.save(commit=False)
            question.section = section
            formset = OptionFormSet(request.POST, instance=question)
            if formset.is_valid():
                question.save()
                formset.save()
                messages.success(request, "Question and options added successfully!")
                return redirect('exam_detail', exam_id=exam.id)
        else:
            formset = OptionFormSet(request.POST)
    else:
        initial_order = section.questions.count() + 1
        form = QuestionForm(initial={'order': initial_order})
        formset = OptionFormSet()
    
    return render(request, 'admin/question_form.html', {
        'form': form,
        'formset': formset,
        'section': section,
        'exam': exam,
        'title': f"Add Question to Section: {section.name}"
    })


@login_required
def question_edit(request, question_id):
    """
    Edits a question (restricted to owner or creator).
    """
    question = get_object_or_404(Question, id=question_id)
    section = question.section
    exam = section.exam
    if not request.user.is_superuser and exam.created_by != request.user:
        raise PermissionDenied
        
    if request.method == 'POST':
        form = QuestionForm(request.POST, request.FILES, instance=question)
        formset = OptionFormSet(request.POST, instance=question)
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            messages.success(request, "Question and options updated successfully!")
            return redirect('exam_detail', exam_id=exam.id)
    else:
        form = QuestionForm(instance=question)
        formset = OptionFormSet(instance=question)
        
    return render(request, 'admin/question_form.html', {
        'form': form,
        'formset': formset,
        'section': section,
        'exam': exam,
        'title': f"Edit Question"
    })


@login_required
def question_delete(request, question_id):
    """
    Deletes a question (restricted to owner or creator).
    """
    question = get_object_or_404(Question, id=question_id)
    exam = question.section.exam
    if not request.user.is_superuser and exam.created_by != request.user:
        raise PermissionDenied
        
    if request.method == 'POST':
        question.delete()
        messages.success(request, "Question deleted successfully.")
        return redirect('exam_detail', exam_id=exam.id)
    return render(request, 'admin/confirm_delete.html', {'object': question, 'back_url': redirect('exam_detail', exam_id=exam.id).url})


# --- EXPORT & IMPORT UTILITIES ---

def export_exam_json(exam):
    data = {
        'title': exam.title,
        'description': exam.description or '',
        'sections': []
    }
    
    for section in exam.sections.all().order_by('order', 'id'):
        sec_data = {
            'name': section.name,
            'description': section.description or '',
            'duration_minutes': section.duration_minutes,
            'duration_seconds': section.duration_seconds,
            'order': section.order,
            'questions': []
        }
        
        for question in section.questions.all().order_by('order', 'id'):
            q_data = {
                'text': question.text,
                'question_type': question.question_type,
                'order': question.order,
                'image': str(question.image) if question.image else None,
                'options': []
            }
            
            for option in question.options.all():
                q_data['options'].append({
                    'text': option.text,
                    'image': str(option.image) if option.image else None,
                    'score': option.score
                })
                
            sec_data['questions'].append(q_data)
            
        data['sections'].append(sec_data)
        
    return data


def export_exam_excel(exam):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Exam Structure"
    
    # Headers
    headers = [
        "Exam Title", "Exam Description", 
        "Section Name", "Section Description", "Section Duration (Minutes)", "Section Duration (Seconds)", "Section Order",
        "Question Text", "Question Type", "Question Order", "Question Image Path",
        "Option Text", "Option Image Path", "Option Score"
    ]
    
    # Apply styling to headers
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    
    row_idx = 2
    sections = exam.sections.all().order_by('order', 'id')
    if not sections.exists():
        ws.cell(row=row_idx, column=1, value=exam.title)
        ws.cell(row=row_idx, column=2, value=exam.description or "")
        row_idx += 1
    else:
        for section in sections:
            questions = section.questions.all().order_by('order', 'id')
            if not questions.exists():
                ws.cell(row=row_idx, column=1, value=exam.title)
                ws.cell(row=row_idx, column=2, value=exam.description or "")
                ws.cell(row=row_idx, column=3, value=section.name)
                ws.cell(row=row_idx, column=4, value=section.description or "")
                ws.cell(row=row_idx, column=5, value=section.duration_minutes)
                ws.cell(row=row_idx, column=6, value=section.duration_seconds)
                ws.cell(row=row_idx, column=7, value=section.order)
                row_idx += 1
            else:
                for question in questions:
                    options = question.options.all()
                    if not options.exists():
                        ws.cell(row=row_idx, column=1, value=exam.title)
                        ws.cell(row=row_idx, column=2, value=exam.description or "")
                        ws.cell(row=row_idx, column=3, value=section.name)
                        ws.cell(row=row_idx, column=4, value=section.description or "")
                        ws.cell(row=row_idx, column=5, value=section.duration_minutes)
                        ws.cell(row=row_idx, column=6, value=section.duration_seconds)
                        ws.cell(row=row_idx, column=7, value=section.order)
                        ws.cell(row=row_idx, column=8, value=question.text)
                        ws.cell(row=row_idx, column=9, value=question.question_type)
                        ws.cell(row=row_idx, column=10, value=question.order)
                        ws.cell(row=row_idx, column=11, value=str(question.image) if question.image else "")
                        row_idx += 1
                    else:
                        for option in options:
                            ws.cell(row=row_idx, column=1, value=exam.title)
                            ws.cell(row=row_idx, column=2, value=exam.description or "")
                            ws.cell(row=row_idx, column=3, value=section.name)
                            ws.cell(row=row_idx, column=4, value=section.description or "")
                            ws.cell(row=row_idx, column=5, value=section.duration_minutes)
                            ws.cell(row=row_idx, column=6, value=section.duration_seconds)
                            ws.cell(row=row_idx, column=7, value=section.order)
                            ws.cell(row=row_idx, column=8, value=question.text)
                            ws.cell(row=row_idx, column=9, value=question.question_type)
                            ws.cell(row=row_idx, column=10, value=question.order)
                            ws.cell(row=row_idx, column=11, value=str(question.image) if question.image else "")
                            ws.cell(row=row_idx, column=12, value=option.text)
                            ws.cell(row=row_idx, column=13, value=str(option.image) if option.image else "")
                            ws.cell(row=row_idx, column=14, value=option.score)
                            row_idx += 1
                            
    # Auto-adjust column widths for readability
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max(max_len + 2, 12)
        
    return wb


def import_exam_json(file_obj):
    try:
        data = json.load(file_obj)
    except Exception as e:
        raise ValueError(f"Invalid JSON file format: {str(e)}")
    
    if 'title' not in data:
        raise ValueError("JSON file is missing 'title' field.")
        
    parsed_sections = {}
    
    sections_list = data.get('sections', [])
    for sec_idx, sec in enumerate(sections_list):
        sec_name = sec.get('name')
        if not sec_name:
            continue
        sec_order = sec.get('order', sec_idx + 1)
        sec_key = (sec_name, sec_order)
        
        parsed_sections[sec_key] = {
            'name': sec_name,
            'description': sec.get('description', ''),
            'duration_minutes': sec.get('duration_minutes', 10),
            'duration_seconds': sec.get('duration_seconds', 0),
            'order': sec_order,
            'questions': {}
        }
        
        questions_list = sec.get('questions', [])
        for q_idx, q in enumerate(questions_list):
            q_text = q.get('text')
            if not q_text:
                continue
            q_order = q.get('order', q_idx + 1)
            q_key = (q_text, q_order)
            
            parsed_sections[sec_key]['questions'][q_key] = {
                'text': q_text,
                'question_type': q.get('question_type', 'single_select'),
                'order': q_order,
                'image': q.get('image'),
                'options': []
            }
            
            options_list = q.get('options', [])
            for opt in options_list:
                opt_text = opt.get('text')
                if opt_text is not None:
                    parsed_sections[sec_key]['questions'][q_key]['options'].append({
                        'text': opt_text,
                        'image': opt.get('image'),
                        'score': float(opt.get('score', 0.0))
                    })
                    
    return {
        'title': data['title'],
        'description': data.get('description', ''),
        'sections': parsed_sections
    }


def import_exam_excel(file_obj):
    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
    except Exception as e:
        raise ValueError(f"Invalid Excel file format: {str(e)}")
        
    ws = wb.active
    
    sections_dict = {}
    exam_title = None
    exam_desc = None
    
    for row in range(2, ws.max_row + 1):
        val = lambda col: ws.cell(row=row, column=col).value
        
        row_exam_title = val(1)
        row_exam_desc = val(2)
        sec_name = val(3)
        sec_desc = val(4)
        sec_dur_min = val(5)
        sec_dur_sec = val(6)
        sec_order = val(7)
        q_text = val(8)
        q_type = val(9)
        q_order = val(10)
        q_image = val(11)
        opt_text = val(12)
        opt_image = val(13)
        opt_score = val(14)
        
        if not exam_title and row_exam_title:
            exam_title = str(row_exam_title).strip()
            if row_exam_desc:
                exam_desc = str(row_exam_desc).strip()
                
        if not sec_name:
            continue
            
        sec_name = str(sec_name).strip()
        try:
            sec_order = int(sec_order) if sec_order is not None else len(sections_dict) + 1
        except ValueError:
            sec_order = len(sections_dict) + 1
            
        sec_key = (sec_name, sec_order)
        
        if sec_key not in sections_dict:
            sections_dict[sec_key] = {
                'name': sec_name,
                'description': str(sec_desc).strip() if sec_desc else "",
                'duration_minutes': int(sec_dur_min) if sec_dur_min is not None else 10,
                'duration_seconds': int(sec_dur_sec) if sec_dur_sec is not None else 0,
                'order': sec_order,
                'questions': {}
            }
            
        sec_data = sections_dict[sec_key]
        
        if q_text:
            q_text = str(q_text).strip()
            try:
                q_order = int(q_order) if q_order is not None else len(sec_data['questions']) + 1
            except ValueError:
                q_order = len(sec_data['questions']) + 1
                
            q_key = (q_text, q_order)
            
            if q_key not in sec_data['questions']:
                sec_data['questions'][q_key] = {
                    'text': q_text,
                    'question_type': str(q_type).strip() if q_type else 'single_select',
                    'order': q_order,
                    'image': str(q_image).strip() if q_image else None,
                    'options': []
                }
                
            q_data = sec_data['questions'][q_key]
            
            if opt_text is not None:
                try:
                    opt_score = float(opt_score) if opt_score is not None else 0.0
                except ValueError:
                    opt_score = 0.0
                q_data['options'].append({
                    'text': str(opt_text).strip(),
                    'image': str(opt_image).strip() if opt_image else None,
                    'score': opt_score
                })
                
    if not exam_title:
        raise ValueError("Excel file is missing an Exam Title in the first column.")
        
    return {
        'title': exam_title,
        'description': exam_desc,
        'sections': sections_dict
    }


def save_parsed_exam_data(parsed_data, exam=None, user=None):
    with transaction.atomic():
        if exam is None:
            exam = Exam(
                title=parsed_data['title'],
                description=parsed_data['description'],
                created_by=user
            )
            exam.save()
        else:
            exam.title = parsed_data['title']
            exam.description = parsed_data['description']
            exam.save()
            exam.sections.all().delete()
            
        for (sec_name, sec_order), sec_val in parsed_data['sections'].items():
            section = Section(
                exam=exam,
                name=sec_val['name'],
                description=sec_val['description'],
                duration_minutes=sec_val['duration_minutes'],
                duration_seconds=sec_val['duration_seconds'],
                order=sec_val['order']
            )
            section.save()
            
            for (q_text, q_order), q_val in sec_val['questions'].items():
                question = Question(
                    section=section,
                    text=q_val['text'],
                    question_type=q_val['question_type'],
                    order=q_val['order'],
                    image=q_val['image'] if q_val['image'] else None
                )
                question.save()
                
                for opt_val in q_val['options']:
                    option = Option(
                        question=question,
                        text=opt_val['text'],
                        image=opt_val.get('image'),
                        score=opt_val['score']
                    )
                    option.save()
        return exam



# --- EXPORT & IMPORT VIEWS ---

@login_required
def exam_export_json(request, exam_id):
    """
    Downloads an exam structure as a JSON file.
    """
    exam = get_object_or_404(Exam, id=exam_id)
    if not request.user.is_superuser and exam.created_by != request.user:
        raise PermissionDenied
        
    data = export_exam_json(exam)
    
    response = HttpResponse(
        json.dumps(data, indent=2),
        content_type='application/json'
    )
    safe_title = "".join(c for c in exam.title if c.isalnum() or c in (' ', '_', '-')).rstrip()
    safe_title = safe_title.replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename="{safe_title}_export.json"'
    return response


@login_required
def exam_export_excel(request, exam_id):
    """
    Downloads an exam structure as an Excel (.xlsx) file.
    """
    exam = get_object_or_404(Exam, id=exam_id)
    if not request.user.is_superuser and exam.created_by != request.user:
        raise PermissionDenied
        
    wb = export_exam_excel(exam)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    safe_title = "".join(c for c in exam.title if c.isalnum() or c in (' ', '_', '-')).rstrip()
    safe_title = safe_title.replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename="{safe_title}_export.xlsx"'
    wb.save(response)
    return response


@login_required
def exam_import(request, exam_id=None):
    """
    Imports an exam from JSON or Excel.
    If exam_id is provided, overwrites the existing exam (creator or superuser only).
    Otherwise, creates a new exam.
    """
    exam = None
    if exam_id:
        exam = get_object_or_404(Exam, id=exam_id)
        if not request.user.is_superuser and exam.created_by != request.user:
            raise PermissionDenied
            
    if request.method == 'POST':
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            messages.error(request, "No file uploaded. Please select a valid JSON or Excel file.")
            return redirect(request.path)
            
        filename = uploaded_file.name.lower()
        try:
            if filename.endswith('.json'):
                parsed_data = import_exam_json(uploaded_file)
            elif filename.endswith('.xlsx'):
                parsed_data = import_exam_excel(uploaded_file)
            else:
                messages.error(request, "Unsupported file format. Please upload a .json or .xlsx file.")
                return render(request, 'admin/exam_import.html', {'exam': exam})
                
            imported_exam = save_parsed_exam_data(parsed_data, exam=exam, user=request.user)
            
            if exam_id:
                messages.success(request, f"Exam '{imported_exam.title}' overwritten and updated successfully!")
            else:
                messages.success(request, f"Exam '{imported_exam.title}' imported and created successfully!")
                
            return redirect('exam_detail', exam_id=imported_exam.id)
            
        except Exception as e:
            messages.error(request, f"Failed to import exam: {str(e)}")
            return render(request, 'admin/exam_import.html', {'exam': exam})
            
    return render(request, 'admin/exam_import.html', {'exam': exam})

