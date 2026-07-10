from django.test import TestCase
from django.utils import timezone
from .models import Exam, Section, Question, Option
from candidates.models import Candidate, ExamSession, CandidateAnswer, ExamAssignment
from candidates.utils import calculate_and_finalize_results

class PsychologicalExamTestCase(TestCase):
    def setUp(self):
        # Create an Exam
        self.exam = Exam.objects.create(
            title="Standard Psychometric Exam",
            description="Tests general logical reasoning and cognitive speed."
        )

        # Create Section
        self.section = Section.objects.create(
            exam=self.exam,
            name="Logic Section",
            duration_minutes=1,
            duration_seconds=0,
            order=1
        )

        # 1. Single Select Question (Max score should be 5.0)
        self.q1 = Question.objects.create(
            section=self.section,
            text="How do you handle stress?",
            question_type="single_select",
            order=1
        )
        self.opt_q1_a = Option.objects.create(question=self.q1, text="Calmly analyze", score=5.0)
        self.opt_q1_b = Option.objects.create(question=self.q1, text="Panic slightly", score=2.0)
        self.opt_q1_c = Option.objects.create(question=self.q1, text="Ignore it", score=0.0)

        # 2. Multi Select Question (Max score should be sum of positive scores = 3.0 + 4.0 = 7.0)
        self.q2 = Question.objects.create(
            section=self.section,
            text="Choose your strengths (select all that apply):",
            question_type="multi_select",
            order=2
        )
        self.opt_q2_a = Option.objects.create(question=self.q2, text="Time management", score=3.0)
        self.opt_q2_b = Option.objects.create(question=self.q2, text="Empathy", score=4.0)
        self.opt_q2_c = Option.objects.create(question=self.q2, text="Procrastination", score=-2.0)

    def test_max_possible_scores(self):
        """
        Verifies max score is calculated correctly for single-select (max option score)
        and multi-select (sum of positive option scores).
        """
        self.assertEqual(self.q1.max_possible_score, 5.0)
        self.assertEqual(self.q2.max_possible_score, 7.0)
        self.assertEqual(self.exam.total_questions, 2)
        self.assertEqual(self.exam.total_sections, 1)

    def test_score_calculation(self):
        """
        Verifies grading and percentages.
        """
        # Create Candidate
        candidate = Candidate.objects.create(
            full_name="John Doe",
            email="john@doe.com",
            mobile_number="1234567890"
        )
        
        # Create Assignment
        assignment = ExamAssignment.objects.create(
            exam=self.exam,
            exam_code="TEST-CODE-1",
            assigned_email="john@doe.com"
        )

        # Create Session
        session = ExamSession.objects.create(
            candidate=candidate,
            assignment=assignment,
            exam=self.exam,
            status="in_progress"
        )

        # John answers Q1: selects "Panic slightly" (+2.0 points)
        ans1 = CandidateAnswer.objects.create(session=session, question=self.q1)
        ans1.selected_options.add(self.opt_q1_b)

        # John answers Q2: selects "Time management" (+3.0 points) and "Procrastination" (-2.0 points)
        # Total points earned = 2.0 (from Q1) + (3.0 - 2.0) = 3.0 points
        # Max possible points = 5.0 + 7.0 = 12.0 points
        # Expected percentage = 3.0 / 12.0 * 100 = 25.0%
        ans2 = CandidateAnswer.objects.create(session=session, question=self.q2)
        ans2.selected_options.add(self.opt_q2_a, self.opt_q2_c)

        # Finalize and calculate
        result = calculate_and_finalize_results(session)

        self.assertEqual(result.overall_score_percentage, 25.0)
        
        # Check section score is recorded
        sec_result = result.section_results.first()
        self.assertIsNotNone(sec_result)
        self.assertEqual(sec_result.score_percentage, 25.0)


from django.contrib.auth.models import User
from django.urls import reverse
from reports.forms import AssignmentForm

class FranchisePermissionsTestCase(TestCase):
    def setUp(self):
        # Create Owner
        self.owner = User.objects.create_superuser(username="owner", email="owner@test.com", password="password123")
        
        # Create Franchisees
        self.franchise1 = User.objects.create_user(username="franchise1", email="f1@test.com", password="password123")
        self.franchise2 = User.objects.create_user(username="franchise2", email="f2@test.com", password="password123")

        # Owner creates exams
        self.owner_private_exam = Exam.objects.create(
            title="Owner Private Exam",
            created_by=self.owner
        )
        self.owner_shared_exam = Exam.objects.create(
            title="Owner Shared Exam",
            created_by=self.owner
        )
        self.owner_shared_exam.shared_with.add(self.franchise1)

        # Franchisee 1 creates their own exam
        self.f1_exam = Exam.objects.create(
            title="Franchise 1 Exam",
            created_by=self.franchise1
        )

    def test_exam_list_permissions(self):
        """
        Franchisees should see only exams they created.
        """
        # Log in as Franchise 1
        self.client.login(username="franchise1", password="password123")
        response = self.client.get(reverse('exam_list'))
        exams = response.context['exams']
        
        # Franchise 1 should see only their own exam.
        self.assertIn(self.f1_exam, exams)
        self.assertNotIn(self.owner_shared_exam, exams)
        self.assertNotIn(self.owner_private_exam, exams)

        # Log in as Franchise 2
        self.client.login(username="franchise2", password="password123")
        response = self.client.get(reverse('exam_list'))
        exams = response.context['exams']
        
        # Franchise 2 should see: None (they created none and none are shared with them)
        self.assertEqual(len(exams), 0)

    def test_exam_detail_and_edit_permissions(self):
        """
        Franchisee can view and edit their own exams only.
        """
        self.client.login(username="franchise1", password="password123")

        # Franchise 1 cannot view or edit an owner/admin exam, even if shared.
        response = self.client.get(reverse('exam_detail', args=[self.owner_shared_exam.id]))
        self.assertEqual(response.status_code, 403)

        response = self.client.get(reverse('exam_edit', args=[self.owner_shared_exam.id]))
        self.assertEqual(response.status_code, 403)

        response = self.client.post(reverse('exam_edit', args=[self.owner_shared_exam.id]), {'title': 'Hacked Title'})
        self.assertEqual(response.status_code, 403)

        # Franchise 1 tries to edit their own exam -> allowed
        response = self.client.get(reverse('exam_edit', args=[self.f1_exam.id]))
        self.assertEqual(response.status_code, 200)

        # Franchise 1 tries to view Owner Private Exam -> Forbidden
        response = self.client.get(reverse('exam_detail', args=[self.owner_private_exam.id]))
        self.assertEqual(response.status_code, 403)

    def test_franchise_can_create_exam(self):
        """
        Franchise users can create exams, and created exams belong to that user.
        """
        self.client.login(username="franchise1", password="password123")
        response = self.client.post(reverse('exam_create'), {
            'title': 'Franchise Created Exam',
            'description': 'Created by franchise user.'
        })

        created_exam = Exam.objects.get(title='Franchise Created Exam')
        self.assertEqual(response.status_code, 302)
        self.assertEqual(created_exam.created_by, self.franchise1)
        self.assertEqual(created_exam.shared_with.count(), 0)

    def test_franchise_assignment_form_only_shows_own_exams(self):
        """
        Franchise users cannot assign candidates to owner/admin exams or other
        franchise exams.
        """
        form = AssignmentForm(user=self.franchise1)
        exams = list(form.fields['exam'].queryset)

        self.assertIn(self.f1_exam, exams)
        self.assertNotIn(self.owner_private_exam, exams)
        self.assertNotIn(self.owner_shared_exam, exams)

    def test_assignment_creation_does_not_send_exam_code_email(self):
        """
        New exam assignments create access records without emailing candidates.
        """
        self.client.login(username="franchise1", password="password123")
        response = self.client.post(reverse('admin_assignments'), {
            'exam': self.f1_exam.id,
            'exam_code': 'F1-CODE-1',
            'assigned_emails': 'candidate@test.com',
        })

        self.assertEqual(response.status_code, 302)
        self.assertTrue(ExamAssignment.objects.filter(
            exam=self.f1_exam,
            exam_code='F1-CODE-1',
            assigned_email='candidate@test.com',
            created_by=self.franchise1,
        ).exists())


import io
import json
import openpyxl
from tests.models import Section, Question, Option

class ExamImportExportTestCase(TestCase):
    def setUp(self):
        # Create Owner and Franchisee
        self.owner = User.objects.create_superuser(username="owner", email="owner@test.com", password="password123")
        self.franchise = User.objects.create_user(username="franchise1", email="f1@test.com", password="password123")
        self.other_franchise = User.objects.create_user(username="franchise2", email="f2@test.com", password="password123")
        
        # Create an exam created by franchise
        self.franchise_exam = Exam.objects.create(
            title="Franchise Cognitive Test",
            description="Cognitive reasoning exam.",
            created_by=self.franchise
        )
        self.sec = Section.objects.create(
            exam=self.franchise_exam,
            name="Section A",
            description="Instructions A",
            duration_minutes=15,
            duration_seconds=30,
            order=1
        )
        self.q = Question.objects.create(
            section=self.sec,
            text="Is 2+2=4?",
            question_type="single_select",
            order=1
        )
        self.opt_yes = Option.objects.create(question=self.q, text="Yes", image="option_images/yes_helper.png", score=10.0)
        self.opt_no = Option.objects.create(question=self.q, text="No", score=0.0)

    def test_export_json_permissions_and_content(self):
        """
        Tests JSON export permissions (owner/creator allowed, other franchisee forbidden) and content structure.
        """
        # Forbidden for other franchise
        self.client.login(username="franchise2", password="password123")
        response = self.client.get(reverse('exam_export_json', args=[self.franchise_exam.id]))
        self.assertEqual(response.status_code, 403)
        
        # Allowed for owner
        self.client.login(username="owner", password="password123")
        response = self.client.get(reverse('exam_export_json', args=[self.franchise_exam.id]))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/json')
        
        # Validate JSON content
        data = json.loads(response.content)
        self.assertEqual(data['title'], "Franchise Cognitive Test")
        self.assertEqual(data['description'], "Cognitive reasoning exam.")
        self.assertEqual(len(data['sections']), 1)
        
        sec_data = data['sections'][0]
        self.assertEqual(sec_data['name'], "Section A")
        self.assertEqual(sec_data['duration_minutes'], 15)
        self.assertEqual(sec_data['duration_seconds'], 30)
        self.assertEqual(len(sec_data['questions']), 1)
        
        q_data = sec_data['questions'][0]
        self.assertEqual(q_data['text'], "Is 2+2=4?")
        self.assertEqual(q_data['question_type'], "single_select")
        self.assertEqual(len(q_data['options']), 2)
        
        self.assertEqual(q_data['options'][0]['text'], "Yes")
        self.assertEqual(q_data['options'][0]['image'], "option_images/yes_helper.png")
        self.assertEqual(q_data['options'][0]['score'], 10.0)

    def test_export_excel_permissions_and_content(self):
        """
        Tests Excel export permissions and content structure.
        """
        # Allowed for creator
        self.client.login(username="franchise1", password="password123")
        response = self.client.get(reverse('exam_export_excel', args=[self.franchise_exam.id]))
        self.assertEqual(response.status_code, 200)
        self.assertIn('spreadsheetml', response['Content-Type'])
        
        # Read excel workbook from response content
        wb = openpyxl.load_workbook(io.BytesIO(response.content), data_only=True)
        ws = wb.active
        self.assertEqual(ws.title, "Exam Structure")
        
        # Check header
        self.assertEqual(ws.cell(row=1, column=1).value, "Exam Title")
        self.assertEqual(ws.cell(row=1, column=3).value, "Section Name")
        
        # Check row data
        self.assertEqual(ws.cell(row=2, column=1).value, "Franchise Cognitive Test")
        self.assertEqual(ws.cell(row=2, column=3).value, "Section A")
        self.assertEqual(ws.cell(row=2, column=5).value, 15)
        self.assertEqual(ws.cell(row=2, column=8).value, "Is 2+2=4?")
        self.assertEqual(ws.cell(row=2, column=12).value, "Yes")
        self.assertEqual(ws.cell(row=2, column=13).value, "option_images/yes_helper.png")
        self.assertEqual(ws.cell(row=2, column=14).value, 10.0)

    def test_import_json_create_new(self):
        """
        Tests importing a JSON file to create a brand new exam.
        """
        self.client.login(username="franchise1", password="password123")
        
        json_data = {
            "title": "Imported Test Exam",
            "description": "JSON Import Description",
            "sections": [
                {
                    "name": "Sec 1",
                    "description": "Sec 1 Instructions",
                    "duration_minutes": 5,
                    "duration_seconds": 0,
                    "order": 1,
                    "questions": [
                        {
                            "text": "Choose true options:",
                            "question_type": "multi_select",
                            "order": 1,
                            "options": [
                                {"text": "Opt A", "image": "option_images/a.png", "score": 2.0},
                                {"text": "Opt B", "score": -1.0}
                            ]
                        }
                    ]
                }
            ]
        }
        
        file_content = json.dumps(json_data).encode('utf-8')
        fp = io.BytesIO(file_content)
        fp.name = "imported_exam.json"
        
        response = self.client.post(reverse('exam_import'), {'file': fp})
        self.assertEqual(response.status_code, 302) # Redirect to detail page
        
        # Verify it created a new exam in database
        new_exam = Exam.objects.get(title="Imported Test Exam")
        self.assertEqual(new_exam.description, "JSON Import Description")
        self.assertEqual(new_exam.created_by, self.franchise)
        self.assertEqual(new_exam.sections.count(), 1)
        
        section = new_exam.sections.first()
        self.assertEqual(section.name, "Sec 1")
        self.assertEqual(section.duration_minutes, 5)
        self.assertEqual(section.questions.count(), 1)
        
        question = section.questions.first()
        self.assertEqual(question.text, "Choose true options:")
        self.assertEqual(question.question_type, "multi_select")
        self.assertEqual(question.options.count(), 2)
        
        opt1 = question.options.get(text="Opt A")
        self.assertEqual(opt1.image, "option_images/a.png")
        self.assertEqual(opt1.score, 2.0)

    def test_import_json_overwrite_existing(self):
        """
        Tests importing a JSON file to overwrite/replace an existing exam.
        """
        self.client.login(username="franchise1", password="password123")
        
        json_data = {
            "title": "Overwritten Title",
            "description": "Overwritten Description",
            "sections": [
                {
                    "name": "Overwritten Sec",
                    "description": "Overwritten Sec Instructions",
                    "duration_minutes": 20,
                    "duration_seconds": 0,
                    "order": 1,
                    "questions": [
                        {
                            "text": "Overwritten question?",
                            "question_type": "single_select",
                            "order": 1,
                            "options": [
                                {"text": "Option 1", "score": 5.0}
                            ]
                        }
                    ]
                }
            ]
        }
        
        file_content = json.dumps(json_data).encode('utf-8')
        fp = io.BytesIO(file_content)
        fp.name = "overwritten_exam.json"
        
        response = self.client.post(reverse('exam_import_overwrite', args=[self.franchise_exam.id]), {'file': fp})
        self.assertEqual(response.status_code, 302)
        
        # Verify it updated the existing exam instead of creating a new one
        self.franchise_exam.refresh_from_db()
        self.assertEqual(self.franchise_exam.title, "Overwritten Title")
        self.assertEqual(self.franchise_exam.description, "Overwritten Description")
        self.assertEqual(self.franchise_exam.sections.count(), 1)
        
        section = self.franchise_exam.sections.first()
        self.assertEqual(section.name, "Overwritten Sec")
        self.assertEqual(section.duration_minutes, 20)
        self.assertEqual(section.questions.count(), 1)
        
        question = section.questions.first()
        self.assertEqual(question.text, "Overwritten question?")
        self.assertEqual(question.options.count(), 1)
        self.assertEqual(question.options.first().score, 5.0)

    def test_import_excel_create_new(self):
        """
        Tests importing an Excel file to create a brand new exam.
        """
        self.client.login(username="franchise1", password="password123")
        
        # Create Excel Workbook in memory
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Exam Structure"
        
        headers = [
            "Exam Title", "Exam Description", 
            "Section Name", "Section Description", "Section Duration (Minutes)", "Section Duration (Seconds)", "Section Order",
            "Question Text", "Question Type", "Question Order", "Question Image Path",
            "Option Text", "Option Image Path", "Option Score"
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
            
        # Add row 1 (Option A)
        ws.cell(row=2, column=1, value="Excel Exam")
        ws.cell(row=2, column=2, value="Excel Description")
        ws.cell(row=2, column=3, value="Excel Section")
        ws.cell(row=2, column=4, value="")
        ws.cell(row=2, column=5, value=12)
        ws.cell(row=2, column=6, value=0)
        ws.cell(row=2, column=7, value=1)
        ws.cell(row=2, column=8, value="Excel Question")
        ws.cell(row=2, column=9, value="single_select")
        ws.cell(row=2, column=10, value=1)
        ws.cell(row=2, column=11, value="")
        ws.cell(row=2, column=12, value="Option A")
        ws.cell(row=2, column=13, value="option_images/opt_a.png")
        ws.cell(row=2, column=14, value=4.5)
        
        # Add row 2 (Option B)
        ws.cell(row=3, column=1, value="Excel Exam")
        ws.cell(row=3, column=2, value="Excel Description")
        ws.cell(row=3, column=3, value="Excel Section")
        ws.cell(row=3, column=4, value="")
        ws.cell(row=3, column=5, value=12)
        ws.cell(row=3, column=6, value=0)
        ws.cell(row=3, column=7, value=1)
        ws.cell(row=3, column=8, value="Excel Question")
        ws.cell(row=3, column=9, value="single_select")
        ws.cell(row=3, column=10, value=1)
        ws.cell(row=3, column=11, value="")
        ws.cell(row=3, column=12, value="Option B")
        ws.cell(row=3, column=13, value="")
        ws.cell(row=3, column=14, value=0.0)
        
        fp = io.BytesIO()
        wb.save(fp)
        fp.seek(0)
        fp.name = "excel_exam.xlsx"
        
        response = self.client.post(reverse('exam_import'), {'file': fp})
        self.assertEqual(response.status_code, 302)
        
        new_exam = Exam.objects.get(title="Excel Exam")
        self.assertEqual(new_exam.description, "Excel Description")
        self.assertEqual(new_exam.sections.count(), 1)
        
        sec = new_exam.sections.first()
        self.assertEqual(sec.name, "Excel Section")
        self.assertEqual(sec.duration_minutes, 12)
        self.assertEqual(sec.questions.count(), 1)
        
        q = sec.questions.first()
        self.assertEqual(q.text, "Excel Question")
        self.assertEqual(q.question_type, "single_select")
        self.assertEqual(q.options.count(), 2)
        opt_a = q.options.get(text="Option A")
        self.assertEqual(opt_a.image, "option_images/opt_a.png")
        self.assertEqual(opt_a.score, 4.5)

